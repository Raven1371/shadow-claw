"""Optional Excel workbook export (``--excel``), using openpyxl."""

from __future__ import annotations

import datetime
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from . import DISCLAIMER, TOOL_NAME, __version__
from .output_safety import OutputSafetyError, atomic_replace_file, create_secure_temp_file
from .sanitization import sanitize_spreadsheet_value
from .models import (
    EvidenceClass,
    FirewallRule,
    Flow,
    Host,
    ManualReviewItem,
    RISK_ORDER,
    ScanMetadata,
    ServiceRecord,
    VulnFinding,
    as_flat_dict,
)

LOG = logging.getLogger("nmap_flow_analyzer.excel_export")

MISSING_OPENPYXL = (
    "openpyxl is not installed, so firewall_exceptions.xlsx was skipped. "
    "Install it with: pip install openpyxl"
)


def _write_sheet(workbook, title: str, columns: Sequence[Tuple[str, str]],
                 rows: Sequence[Dict[str, Any]], sanitize: bool = True) -> None:
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    sheet = workbook.create_sheet(title=title[:31])
    bold = Font(bold=True)
    for col_index, (_, label) in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_index, value=label)
        cell.font = bold
    widths = [len(label) for _, label in columns]
    for row_index, row in enumerate(rows, start=2):
        for col_index, (key, _) in enumerate(columns, start=1):
            value = row.get(key, "")
            if sanitize:
                value = sanitize_spreadsheet_value(value)
            sheet.cell(row=row_index, column=col_index, value=value)
            widths[col_index - 1] = max(widths[col_index - 1], min(len(str(value)), 60))
    for col_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col_index)].width = min(width + 2, 62)
    sheet.freeze_panes = "A2"
    wrap = Alignment(wrap_text=True, vertical="top")
    for row_cells in sheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = wrap


def export_excel(
    path: Path,
    metadata: ScanMetadata,
    hosts: List[Host],
    services: List[ServiceRecord],
    flows: List[Flow],
    inbound: List[FirewallRule],
    outbound: List[FirewallRule],
    reviews: List[ManualReviewItem],
    vulns: List[VulnFinding],
    assumptions: List[str],
    sanitize: bool = True,
    run_status: Optional[Dict[str, str]] = None,
    zeek: Optional[Any] = None,
) -> Optional[str]:
    """Write the workbook; returns an error/skip message or None on success.

    All untrusted string cells are spreadsheet-sanitized unless ``sanitize``
    is False (not recommended; see --no-spreadsheet-sanitization).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        LOG.warning(MISSING_OPENPYXL)
        return MISSING_OPENPYXL

    workbook = openpyxl.Workbook()
    summary = workbook.active
    summary.title = "Executive Summary"
    open_services = [s for s in services if s.state == "open"]
    high = [s for s in open_services if RISK_ORDER[s.risk_level] >= RISK_ORDER["High"]]
    pairs = [
        ("Report", f"{TOOL_NAME} {__version__}"),
        ("Generated", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Input file", metadata.source_file),
        ("Nmap command line", metadata.command_line),
        ("Scan started", metadata.start_time),
        ("Scan finished", metadata.end_time),
        ("Hosts parsed", len(hosts)),
        ("Open services", len(open_services)),
        ("Candidate inbound rules", len(inbound)),
        ("Candidate outbound rules", len(outbound)),
        ("Manual-review items", len(reviews)),
        ("High/Critical exposed services", len(high)),
        ("Disclaimer", DISCLAIMER),
        ("Spreadsheet sanitization",
         "applied" if sanitize else "DISABLED (--no-spreadsheet-sanitization)"),
    ]
    for key, value in (run_status or {}).items():
        pairs.append((key, value))
    bold = Font(bold=True)
    for row_index, (label, value) in enumerate(pairs, start=1):
        summary.cell(row=row_index, column=1, value=label).font = bold
        cell_value = str(value)
        if sanitize:
            cell_value = sanitize_spreadsheet_value(cell_value)
        summary.cell(row=row_index, column=2, value=cell_value)
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 110

    host_rows = [
        {
            "hostname": h.display_name,
            "ip": h.primary_ip,
            "zone": h.zone,
            "role": h.role,
            "role_confidence": h.role_confidence,
            "role_source": h.role_source,
            "os": h.best_os,
            "mac": h.mac,
            "mac_vendor": h.mac_vendor,
            "state": h.state,
            "open_ports": len(h.open_ports()),
            "owner": h.owner,
            "purpose": h.purpose,
            "device_class": h.device_class,
        }
        for h in hosts
    ]
    _write_sheet(
        workbook,
        "Hosts",
        [
            ("hostname", "Hostname"), ("ip", "IP"), ("zone", "Zone"),
            ("role", "Role"), ("role_confidence", "Role conf."),
            ("role_source", "Role source"), ("os", "OS guess"), ("mac", "MAC"),
            ("mac_vendor", "MAC vendor"), ("state", "State"),
            ("open_ports", "Open ports"), ("owner", "Owner"),
            ("purpose", "Purpose"), ("device_class", "Device class"),
        ],
        host_rows,
        sanitize=sanitize
    )
    service_cols = [
        ("hostname", "Host"), ("ip", "IP"), ("protocol", "Protocol"),
        ("port", "Port"), ("state", "State"), ("normalized_service", "Service"),
        ("detected_service", "Detected"), ("expected_service", "Expected"),
        ("mismatch", "Mismatch"), ("product", "Product"), ("version", "Version"),
        ("tunnel", "Tunnel"), ("risk_level", "Risk"),
        ("risk_reasons", "Risk reasons"), ("encryption", "Encryption"),
        ("authentication", "Authentication"), ("evidence_class", "Evidence"),
        ("approved", "Approved"),
    ]
    _write_sheet(workbook, "Services", service_cols, [as_flat_dict(s) for s in services], sanitize=sanitize)

    rule_cols = [
        ("rule_id", "Rule ID"), ("target_hostname", "Target host"),
        ("target_ip", "Target IP"), ("target_zone", "Target zone"),
        ("source", "Source"), ("source_zone", "Source zone"),
        ("destination", "Destination"), ("destination_zone", "Dest. zone"),
        ("protocol", "Protocol"), ("port", "Port"), ("service", "Service"),
        ("purpose", "Business purpose"), ("evidence_class", "Evidence"),
        ("confidence", "Confidence"), ("encryption", "Encryption"),
        ("authentication", "Authentication"), ("risk_level", "Risk"),
        ("recommended_action", "Recommended action"),
        ("description", "Proposed rule"), ("scope_warning", "Scope warning"),
        ("manual_review_notes", "Review notes"),
        ("flow_evidence", "Combined evidence"),
        ("service_evidence", "Service evidence"),
        ("source_scope_evidence", "Source-scope evidence"),
        ("destination_scope_evidence", "Dest-scope evidence"),
        ("perspective", "Perspective"), ("enforcement_model", "Enforcement model"),
        ("source_scope_status", "Source policy status"),
        ("destination_scope_status", "Dest policy status"),
        ("matched_source_policy", "Matched source policy"),
        ("matched_destination_policy", "Matched dest policy"),
        ("policy_violation", "Policy violation"), ("policy_notes", "Policy notes"),
    ]
    _write_sheet(workbook, "Inbound Exceptions", rule_cols, [as_flat_dict(r) for r in inbound], sanitize=sanitize)
    _write_sheet(workbook, "Outbound Exceptions", rule_cols, [as_flat_dict(r) for r in outbound], sanitize=sanitize)
    _write_sheet(
        workbook,
        "Manual Review",
        [
            ("item_id", "ID"), ("category", "Category"), ("host", "Host"),
            ("ip", "IP"), ("protocol", "Protocol"), ("port", "Port"),
            ("finding", "Finding"), ("reason", "Reason"),
            ("recommendation", "Recommendation"), ("related_flow", "Related flow"),
        ],
        [as_flat_dict(m) for m in reviews],
        sanitize=sanitize
    )
    _write_sheet(
        workbook,
        "Vulnerability Findings",
        [
            ("host", "Host"), ("ip", "IP"), ("protocol", "Protocol"),
            ("port", "Port"), ("script_id", "NSE script"),
            ("category", "Category"), ("excerpt", "Output (as reported)"),
        ],
        [as_flat_dict(v) for v in vulns],
        sanitize=sanitize
    )
    _write_sheet(
        workbook,
        "Assumptions",
        [("n", "#"), ("assumption", "Assumption")],
        [{"n": i, "assumption": a} for i, a in enumerate(assumptions, start=1)],
        sanitize=sanitize
    )
    _write_sheet(
        workbook,
        "Evidence",
        [
            ("flow_id", "Flow ID"), ("source", "Source"),
            ("destination", "Destination"), ("protocol", "Protocol"),
            ("destination_port", "Port"), ("normalized_service", "Service"),
            ("evidence_class", "Evidence class"), ("confidence", "Confidence"),
            ("port_state", "Nmap state"), ("state_reason", "State reason"),
            ("evidence", "Evidence description"),
            ("flow_evidence", "Combined evidence"),
            ("perspective", "Perspective"),
            ("related_scripts", "Related NSE scripts"),
        ],
        [as_flat_dict(f) for f in flows],
        sanitize=sanitize
    )
    if zeek is not None:
        agg_cols = [
            ("originator", "Originator"), ("responder", "Responder"),
            ("protocol", "Protocol"), ("responder_port", "Port"),
            ("outcome_label", "Outcome"), ("services", "Zeek services"),
            ("connection_count", "Connections"),
            ("successful_count", "Successful"),
            ("failed_count", "Failed"), ("rejected_count", "Rejected"),
            ("one_way_count", "One-way"), ("partial_count", "Partial"),
            ("first_seen", "First seen (UTC)"), ("last_seen", "Last seen (UTC)"),
            ("originator_bytes", "Orig bytes"),
            ("responder_bytes", "Resp bytes"),
            ("originator_packets", "Orig packets"),
            ("responder_packets", "Resp packets"),
            ("conn_states", "Conn states"), ("missed_bytes", "Missed bytes"),
            ("sample_uids", "Sample UIDs"),
            ("scanner_traffic", "Scanner traffic"),
            ("sensor_quality", "Sensor quality"),
            ("dns_names", "DNS names"), ("tls_server_names", "TLS names"),
            ("http_hosts", "HTTP hosts"),
        ]
        _write_sheet(
            workbook, "Zeek Observed Flows", agg_cols,
            [as_flat_dict(a) for a in zeek.aggregates], sanitize=sanitize,
        )
        _write_sheet(
            workbook, "Correlation Findings",
            [
                ("finding_id", "ID"), ("category", "Category"),
                ("source", "Source"), ("destination", "Destination"),
                ("protocol", "Protocol"), ("port", "Port"),
                ("detail", "Detail"), ("evidence", "Evidence"),
                ("recommendation", "Recommendation"),
                ("correlation_status", "Correlation status"),
            ],
            [as_flat_dict(f) for f in zeek.findings], sanitize=sanitize,
        )
        _write_sheet(
            workbook, "External Dependencies", agg_cols,
            [as_flat_dict(a) for a in zeek.external_dependencies],
            sanitize=sanitize,
        )
        from .zeek_correlation import (
            count_communications, evidence_overlap_breakdown,
            supported_by_nmap_and_zeek,
        )

        corroborated_flows = [
            f for f in flows if supported_by_nmap_and_zeek(f)
        ]
        zeek_only_flows = [
            f for f in flows if f.correlation_status == "zeek-traffic-only"
        ]
        conflicting_flows = [
            f for f in flows if f.correlation_status == "conflicting-evidence"
        ]
        xl_overlap = evidence_overlap_breakdown(flows)
        health = zeek.health
        health_rows = [
            {"key": "Communications supported by both Nmap and Zeek "
                    "(evidence overlap)",
             "value": xl_overlap["supported_by_nmap_and_zeek"]},
            {"key": "  Nmap-and-Zeek only",
             "value": xl_overlap["nmap_and_zeek_only"]},
            {"key": "  Declared and observed with Nmap corroboration",
             "value": xl_overlap["declared_with_nmap_corroboration"]},
            {"key": "  Inferred and observed with Nmap corroboration",
             "value": xl_overlap["inferred_with_nmap_corroboration"]},
            {"key": "Corroborated firewall perspectives",
             "value": len(corroborated_flows)},
            {"key": "Unique Zeek-only communications",
             "value": count_communications(zeek_only_flows)},
            {"key": "Zeek-only firewall perspectives",
             "value": len(zeek_only_flows)},
            {"key": "Unique conflicting-evidence communications (rules withheld)",
             "value": count_communications(conflicting_flows)},
            {"key": "Nmap services with non-scanner production traffic",
             "value": sum(1 for s in services if s.state == "open"
                          and s.non_scanner_zeek_traffic_observed)},
            {"key": "Nmap services with scanner-only traffic",
             "value": sum(1 for s in services if s.state == "open"
                          and s.scanner_only_zeek_traffic_observed)},
            {"key": "Nmap services with no Zeek traffic",
             "value": sum(1 for s in services if s.state == "open"
                          and not s.any_zeek_traffic_observed)},
            {"key": "Status", "value": health.status},
            {"key": "Observation window (UTC)",
             "value": f"{health.observation_start or 'unknown'} to "
                      f"{health.observation_end or 'unknown'}"},
            {"key": "Capture-loss data available",
             "value": "yes" if health.capture_loss_available else "no"},
            {"key": "Max capture loss (%)",
             "value": f"{health.max_capture_loss_percent:.2f}"},
            {"key": "Aggregates with missed bytes",
             "value": health.aggregates_with_missed_bytes},
            {"key": "Total missed bytes", "value": health.total_missed_bytes},
            {"key": "Reporter errors",
             "value": "; ".join(health.reporter_errors) or "none"},
            {"key": "Reporter warnings",
             "value": "; ".join(health.reporter_warnings) or "none"},
            {"key": "Malformed input records",
             "value": health.malformed_records},
        ] + [{"key": "Note", "value": n} for n in health.notes]
        _write_sheet(
            workbook, "Zeek Sensor Health",
            [("key", "Item"), ("value", "Value")], health_rows,
            sanitize=sanitize,
        )
    # Atomic publish: save to a validated temp file in the same directory,
    # verify the workbook reopens cleanly, then os.replace() into place.  A
    # failure never truncates or corrupts an existing workbook.
    tmp = create_secure_temp_file(path.parent, suffix=".xlsx")
    try:
        workbook.save(tmp)
        openpyxl.load_workbook(tmp, read_only=True).close()  # sanity check
        atomic_replace_file(tmp, path)
    except (OSError, OutputSafetyError, Exception) as exc:
        tmp.unlink(missing_ok=True)
        message = f"Could not write {path.name}: {exc}"
        LOG.error(message)
        return message
    LOG.info("Wrote %s", path.name)
    return None
