"""Tests for Zeek-driven rule generation, output safety, and publication."""

from __future__ import annotations

import csv
import json
import tempfile
from pathlib import Path
from unittest import mock

from nmap_flow_analyzer.cli import main
from nmap_flow_analyzer.firewall_rules import build_inbound, build_outbound
from nmap_flow_analyzer.flow_analysis import AnalysisOptions
from tests.fixtures import MULTI_HOST_XML
from tests.zeek_fixtures import conn, correlated, write_json_log, zeek_config


def _rules(data):
    inbound, in_rev = build_inbound(data["flows"], data["config"], data["options"])
    outbound, out_rev = build_outbound(data["flows"], inbound, data["config"],
                                       data["options"])
    return inbound, outbound, in_rev + out_rev


def test_successful_zeek_flow_creates_source_specific_perspectives():
    with tempfile.TemporaryDirectory() as tmp:
        write_json_log(Path(tmp), "conn.log", [
            conn("C1", "192.168.20.15", "192.168.10.30", 5432,
                 service="postgresql"),
        ])
        inbound, outbound, _ = _rules(correlated(Path(tmp)))
        inb = [r for r in inbound if r.destination == "192.168.10.30"
               and r.port == "5432" and r.source == "192.168.20.15"]
        outb = [r for r in outbound if r.destination == "192.168.10.30"
                and r.port == "5432" and r.source == "192.168.20.15"]
        assert inb, "destination-inbound perspective missing"
        assert outb, "source-outbound perspective missing"
        # The actual observed originator, never a widened subnet.
        assert inb[0].source == "192.168.20.15"
        assert inb[0].correlation_status == "nmap-and-zeek"
        assert "zeek" in inb[0].evidence_sources
        assert "authorization" in (inb[0].evidence + inb[0].recommended_action).lower()


def test_network_only_enforcement_creates_single_rule():
    cfg = zeek_config()
    cfg.rule_generation.enforcement_model = "network_only"
    with tempfile.TemporaryDirectory() as tmp:
        write_json_log(Path(tmp), "conn.log", [
            conn("C1", "192.168.20.15", "192.168.10.30", 5432),
        ])
        data = correlated(Path(tmp), cfg)
        zeek_flows = [f for f in data["flows"]
                      if f.source == "192.168.20.15"
                      and f.destination_port == 5432]
        assert len(zeek_flows) == 1
        assert zeek_flows[0].perspective == "network"


def test_perspective_toggles_respected():
    cfg = zeek_config()
    cfg.rule_generation.generate_source_outbound = False
    with tempfile.TemporaryDirectory() as tmp:
        write_json_log(Path(tmp), "conn.log", [
            conn("C1", "192.168.20.15", "192.168.10.30", 5432),
        ])
        data = correlated(Path(tmp), cfg)
        perspectives = {f.perspective for f in data["flows"]
                        if f.source == "192.168.20.15"
                        and f.destination_port == 5432}
        assert perspectives == {"destination-inbound"}


def test_failed_and_one_way_traffic_creates_no_rules():
    with tempfile.TemporaryDirectory() as tmp:
        write_json_log(Path(tmp), "conn.log", [
            conn("C1", "192.168.20.15", "192.168.10.30", 445, state="REJ"),
            conn("C2", "192.168.20.15", "192.168.10.30", 5433, state="S0"),
            conn("C3", "192.168.20.15", "192.168.10.20", 162, proto="udp",
                 state="S0", resp_bytes=0, resp_pkts=0),
            conn("C4", "192.168.20.15", "192.168.10.20", 9099, state="OTH"),
        ])
        inbound, outbound, _ = _rules(correlated(Path(tmp)))
        for rule in inbound + outbound:
            assert (rule.destination, rule.port) not in {
                ("192.168.10.30", "445"), ("192.168.10.30", "5433"),
                ("192.168.10.20", "162"), ("192.168.10.20", "9099"),
            }, f"rule from unsuccessful traffic: {rule.rule_id}"


def test_scanner_traffic_creates_no_production_rule_by_default():
    with tempfile.TemporaryDirectory() as tmp:
        # Scanner hits an un-scanned port; would be a new flow if not excluded.
        write_json_log(Path(tmp), "conn.log", [
            conn("Cs", "192.168.1.50", "192.168.40.9", 8080),
        ])
        inbound, outbound, _ = _rules(correlated(Path(tmp)))
        assert not any(r.destination == "192.168.40.9" for r in inbound + outbound)


def test_ignored_sources_and_destinations_create_no_rules():
    cfg = zeek_config(ignored_sources=["192.168.20.15"],
                      ignored_destinations=["192.168.50.0/24"])
    with tempfile.TemporaryDirectory() as tmp:
        write_json_log(Path(tmp), "conn.log", [
            conn("C1", "192.168.20.15", "192.168.10.30", 5432),
            conn("C2", "192.168.30.4", "192.168.50.9", 8443),
        ])
        data = correlated(Path(tmp), cfg)
        inbound, outbound, _ = _rules(data)
        for rule in inbound + outbound:
            assert rule.source != "192.168.20.15"
            assert not rule.destination.startswith("192.168.50.")
        # Still present in input statistics.
        assert data["zeek"].metadata.record_counts["conn"] == 2


def test_minimum_thresholds_gate_rule_eligibility():
    cfg = zeek_config(minimum_successful_connections_for_rule=3)
    with tempfile.TemporaryDirectory() as tmp:
        write_json_log(Path(tmp), "conn.log", [
            conn("C1", "192.168.20.15", "192.168.10.30", 5432),
            conn("C2", "192.168.20.15", "192.168.10.30", 5432, ts=20.0),
        ])
        data = correlated(Path(tmp), cfg)
        assert not [f for f in data["flows"] if f.source == "192.168.20.15"
                    and f.destination_port == 5432]
        assert data["corr"].attempted  # tracked, not ruled
    cfg = zeek_config(minimum_success_ratio_for_rule=0.9)
    with tempfile.TemporaryDirectory() as tmp:
        write_json_log(Path(tmp), "conn.log", [
            conn("C1", "192.168.20.15", "192.168.10.30", 5432, state="SF"),
            conn("C2", "192.168.20.15", "192.168.10.30", 5432, state="REJ",
                 ts=20.0),
        ])
        data = correlated(Path(tmp), cfg)  # ratio 0.5 < 0.9
        assert not [f for f in data["flows"] if f.source == "192.168.20.15"
                    and f.destination_port == 5432]


def test_policy_and_strict_mode_apply_to_zeek_rules():
    cfg = zeek_config()
    cfg.global_policy = {"approved_source_networks": ["192.168.10.0/24"]}
    options = AnalysisOptions(strict=True)
    with tempfile.TemporaryDirectory() as tmp:
        write_json_log(Path(tmp), "conn.log", [
            conn("C1", "192.168.20.15", "192.168.10.30", 5432),  # outside policy
        ])
        data = correlated(Path(tmp), cfg, options=options)
        inbound, outbound, reviews = _rules(data)
        listed = [r for r in inbound if r.source == "192.168.20.15"
                  and r.port == "5432"]
        withheld = [rv for rv in reviews
                    if rv.ip == "192.168.10.30" and rv.port == "5432"]
        # Strict mode: policy violation is withheld to manual review.
        assert not listed
        assert withheld


def test_stateful_no_return_rule_stateless_keeps_behavior():
    with tempfile.TemporaryDirectory() as tmp:
        write_json_log(Path(tmp), "conn.log", [
            conn("C1", "192.168.20.15", "192.168.10.30", 5432),
        ])
        cfg = zeek_config()
        cfg.firewall_mode = "stateful"
        data = correlated(Path(tmp), cfg,
                          options=AnalysisOptions(firewall_mode="stateful"))
        _, outbound, _ = _rules(data)
        assert not any("return" in r.service.lower() for r in outbound)

        cfg = zeek_config()
        cfg.firewall_mode = "stateless"
        data = correlated(Path(tmp), cfg,
                          options=AnalysisOptions(firewall_mode="stateless"))
        _, outbound, _ = _rules(data)
        returns = [r for r in outbound if "return" in r.service.lower()
                   and r.destination == "192.168.20.15"]
        assert returns  # explicit return path for the Zeek-derived rule


def test_no_any_any_regression():
    with tempfile.TemporaryDirectory() as tmp:
        write_json_log(Path(tmp), "conn.log", [
            conn("C1", "192.168.20.15", "192.168.10.30", 5432),
            conn("C2", "192.168.20.15", "203.0.113.50", 443),
        ])
        cfg = zeek_config(include_external_dependencies_in_rules=True)
        inbound, outbound, _ = _rules(correlated(Path(tmp), cfg))
        for rule in inbound + outbound:
            assert rule.source.lower() not in ("any", "*", "0.0.0.0/0")
            assert rule.destination.lower() not in ("any", "*", "0.0.0.0/0")
            assert rule.port.lower() not in ("any", "*")


# ---------------------------------------------------------------------------
# Full CLI runs: artifacts, publication, cleanup, hostility
# ---------------------------------------------------------------------------

ZEEK_ARTIFACTS = [
    "zeek_observed_flows.csv", "zeek_observed_flows.json",
    "correlation_findings.csv", "correlation_findings.json",
    "external_dependencies.csv", "external_dependencies.json",
    "zeek_input_summary.json",
]


def _write_scan(tmp: Path) -> Path:
    scan = tmp / "scan.xml"
    scan.write_text(MULTI_HOST_XML, encoding="utf-8")
    return scan


def _basic_zeek_dir(tmp: Path) -> Path:
    zdir = tmp / "zeek"
    write_json_log(zdir, "conn.log", [
        conn("C1", "192.168.20.15", "192.168.10.30", 5432,
             service="postgresql"),
        conn("C2", "192.168.20.15", "203.0.113.50", 443, service="ssl"),
        conn("C3", "192.168.20.15", "192.168.10.30", 445, state="REJ"),
    ])
    return zdir


def _run(tmp: Path, out: Path, extra=None) -> int:
    return main([
        "--input", str(tmp / "scan.xml"), "--output-dir", str(out),
        "--scanner-ip", "192.168.1.50",
    ] + (extra or []))


def test_combined_run_generates_and_publishes_zeek_artifacts():
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        _write_scan(tmp)
        zdir = _basic_zeek_dir(tmp)
        out = tmp / "out"
        assert _run(tmp, out, ["--zeek-dir", str(zdir)]) == 0
        for name in ZEEK_ARTIFACTS:
            assert (out / name).is_file(), name
        manifest = json.loads((out / "run_manifest.json").read_text("utf-8"))
        listed = {e["path"] for e in manifest["generated_files"]}
        assert set(ZEEK_ARTIFACTS) <= listed
        zb = manifest["zeek"]
        assert zb["requested"] and zb["active"] and zb["artifacts_generated"]
        assert zb["record_counts"]["conn"] == 3
        assert zb["log_types_found"] == ["conn"]
        assert zb["earliest_timestamp"].startswith("2026-07-20T")
        summary = json.loads((out / "zeek_input_summary.json").read_text("utf-8"))
        assert summary["input_metadata"]["conn_log_available"] is True
        assert summary["sensor_health"]["status"] in (
            "healthy", "degraded", "unknown"
        )
        html = (out / "analysis_report.html").read_text("utf-8")
        assert "Nmap + Zeek Flow Analysis Report" in html
        assert "Not observed during the Zeek collection window" in html
        assert "Zeek sensor health" in html
        dot = (out / "data_flow_diagram.dot").read_text("utf-8")
        assert "Network Data-Flow Diagram (Nmap + Zeek)" in dot
        nd = json.loads((out / "normalized_data.json").read_text("utf-8"))
        for key in ("zeek_input_metadata", "zeek_sensor_health",
                    "zeek_flow_aggregates", "endpoint_identities",
                    "correlation_findings", "external_dependencies"):
            assert key in nd, key
        assert nd["zeek_flow_aggregates"]


def test_nmap_only_run_unchanged_and_without_zeek_artifacts():
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        _write_scan(tmp)
        out = tmp / "out"
        assert _run(tmp, out) == 0
        for name in ZEEK_ARTIFACTS:
            assert not (out / name).exists(), name
        html = (out / "analysis_report.html").read_text("utf-8")
        assert "<h1>Nmap Flow Analysis Report</h1>" in html
        assert "Nmap + Zeek" not in html
        dot = (out / "data_flow_diagram.dot").read_text("utf-8")
        assert "Network Data-Flow Diagram (Nmap-derived)" in dot
        nd = json.loads((out / "normalized_data.json").read_text("utf-8"))
        assert "zeek_flow_aggregates" not in nd
        manifest = json.loads((out / "run_manifest.json").read_text("utf-8"))
        assert manifest["zeek"]["requested"] is False


def test_stale_zeek_artifacts_removed_by_nmap_only_overwrite():
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        _write_scan(tmp)
        zdir = _basic_zeek_dir(tmp)
        out = tmp / "out"
        assert _run(tmp, out, ["--zeek-dir", str(zdir)]) == 0
        keep = out / "administrator-notes.txt"
        keep.write_text("keep me", encoding="utf-8")
        assert _run(tmp, out, ["--overwrite"]) == 0  # Nmap-only
        for name in ZEEK_ARTIFACTS:
            assert not (out / name).exists(), f"stale {name} survived"
        assert keep.read_text(encoding="utf-8") == "keep me"


def test_zeek_required_fatal_when_conn_missing():
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        _write_scan(tmp)
        empty = tmp / "zeek-empty"
        empty.mkdir()
        out = tmp / "out"
        assert _run(tmp, out, ["--zeek-dir", str(empty),
                               "--zeek-required"]) == 2
        assert not (out / "analysis_report.html").exists()
        # Without the flag: warning + Nmap-only completion.
        assert _run(tmp, tmp / "out2", ["--zeek-dir", str(empty)]) == 0
        assert not (tmp / "out2" / "zeek_observed_flows.csv").exists()
        manifest = json.loads(
            (tmp / "out2" / "run_manifest.json").read_text("utf-8")
        )
        assert manifest["zeek"]["requested"] is True
        assert manifest["zeek"]["active"] is False


def test_symlinked_zeek_artifact_destination_rejected():
    import os

    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        _write_scan(tmp)
        zdir = _basic_zeek_dir(tmp)
        out = tmp / "out"
        out.mkdir()
        external = tmp / "external.txt"
        external.write_text("EXTERNAL", encoding="utf-8")
        try:
            os.symlink(external, out / "zeek_observed_flows.csv")
        except (OSError, NotImplementedError):
            return  # platform cannot create symlinks
        code = _run(tmp, out, ["--zeek-dir", str(zdir), "--overwrite"])
        assert code == 2
        assert external.read_text(encoding="utf-8") == "EXTERNAL"


def test_graphviz_failure_in_combined_run_is_nonfatal():
    from nmap_flow_analyzer import diagrams

    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        _write_scan(tmp)
        zdir = _basic_zeek_dir(tmp)
        out = tmp / "out"
        with mock.patch.object(diagrams.shutil, "which", return_value=None):
            assert _run(tmp, out, ["--zeek-dir", str(zdir)]) == 0
        assert (out / "zeek_observed_flows.csv").exists()
        assert not (out / "data_flow_diagram.svg").exists()


def test_fatal_failure_leaves_previous_combined_run_intact():
    import nmap_flow_analyzer.cli as cli

    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        _write_scan(tmp)
        zdir = _basic_zeek_dir(tmp)
        out = tmp / "out"
        assert _run(tmp, out, ["--zeek-dir", str(zdir)]) == 0
        before = {
            e.name: e.read_bytes() for e in out.iterdir()
            if e.is_file() and e.name != "execution.log"
        }
        with mock.patch.object(cli, "write_html_report",
                               side_effect=RuntimeError("boom")):
            assert _run(tmp, out, ["--zeek-dir", str(zdir),
                                   "--overwrite"]) == 1
        after = {
            e.name: e.read_bytes() for e in out.iterdir()
            if e.is_file() and e.name != "execution.log"
        }
        assert after == before
        assert not list(out.glob(".nmap-flow-analyzer-stage-*"))


HOSTILE_SERVICE = "=cmd|'/c calc'!A0"
HOSTILE_SNI = '<script>alert(1)</script>"]; nodeX --> evil["pwn'
HOSTILE_DNS = "=HYPERLINK(\"http://evil\")\n<b>bold</b> `tick` |pipe|"
HOSTILE_LONG = "A" * 3000


def test_hostile_zeek_values_sanitized_in_every_output():
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        _write_scan(tmp)
        zdir = tmp / "zeek"
        write_json_log(zdir, "conn.log", [
            conn("C1", "192.168.20.15", "192.168.10.30", 5432,
                 service=HOSTILE_SERVICE),
        ])
        write_json_log(zdir, "ssl.log", [
            {"ts": 1.0, "uid": "C1", "server_name": HOSTILE_SNI,
             "version": "TLSv1.3", "subject": HOSTILE_LONG},
        ])
        write_json_log(zdir, "dns.log", [
            {"ts": 1.0, "uid": "C1", "query": HOSTILE_DNS,
             "answers": ["192.168.10.30"]},
        ])
        out = tmp / "out"
        assert _run(tmp, out, ["--zeek-dir", str(zdir), "--excel",
                               "--diagram-detail", "full"]) == 0

        # CSV: formula-injection neutralized in the Zeek artifact.
        with (out / "zeek_observed_flows.csv").open(newline="",
                                                    encoding="utf-8") as fh:
            rows = list(csv.DictReader(fh))
        cell = rows[0]["services"]
        assert HOSTILE_SERVICE in cell and cell.startswith("'")
        # HTML: script never appears unescaped anywhere.
        html = (out / "analysis_report.html").read_text("utf-8")
        assert "<script>alert(1)</script>" not in html
        assert "alert(1)" in html  # visible, escaped
        # DOT/Mermaid: hostile SNI cannot break out of labels.
        dot = (out / "data_flow_diagram.dot").read_text("utf-8")
        assert 'evil["pwn' not in dot
        mmd = (out / "data_flow_diagram.mmd").read_text("utf-8")
        assert "-> evil" not in mmd and "--> evil" not in mmd
        # Markdown summary intact (hostile newline collapsed / escaped).
        md = (out / "findings_summary.md").read_text("utf-8")
        assert "<script>alert(1)" not in md
        # Excel written and reopenable despite hostile content.
        import openpyxl
        wb = openpyxl.load_workbook(out / "firewall_exceptions.xlsx",
                                    read_only=True)
        assert "Zeek Observed Flows" in wb.sheetnames
        assert "Correlation Findings" in wb.sheetnames
        assert "External Dependencies" in wb.sheetnames
        assert "Zeek Sensor Health" in wb.sheetnames
        wb.close()


def test_zeek_rules_carry_correlation_fields_in_outputs():
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        _write_scan(tmp)
        zdir = _basic_zeek_dir(tmp)
        out = tmp / "out"
        assert _run(tmp, out, ["--zeek-dir", str(zdir)]) == 0
        rules = json.loads((out / "inbound_exceptions.json").read_text("utf-8"))
        zeek_rules = [r for r in rules if "zeek" in r["evidence_sources"]]
        assert zeek_rules
        rule = next(r for r in zeek_rules if r["source"] == "192.168.20.15")
        assert rule["correlation_status"] == "nmap-and-zeek"
        assert rule["connection_count"] == 1
        assert rule["first_seen"].startswith("2026-07-20T")
