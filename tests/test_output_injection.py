"""End-to-end injection tests: hostile scan values through every writer."""

import csv
import tempfile
from pathlib import Path

from nmap_flow_analyzer.diagrams import build_dot, build_mermaid
from nmap_flow_analyzer.flow_analysis import AnalysisOptions
from nmap_flow_analyzer.models import FirewallRule, ManualReviewItem, ServiceRecord
from nmap_flow_analyzer.reports import (
    write_csv,
    write_findings_summary,
    write_html_report,
)
from tests.fixtures import HOSTILE_XML, base_config, run_pipeline

FORMULA_STARTS = ("=", "+", "-", "@")


def _hostile_run():
    return run_pipeline(HOSTILE_XML, base_config(), AnalysisOptions())


def test_csv_cells_cannot_be_formulas():
    data = _hostile_run()
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "services.csv"
        write_csv(path, data["records"], ServiceRecord)
        with path.open(newline="", encoding="utf-8") as handle:
            rows = list(csv.DictReader(handle))
        assert rows
        for row in rows:
            for value in row.values():
                assert not str(value).startswith(FORMULA_STARTS), value
        # The hostile product string survives, quoted as literal text.
        assert any("+SUM(A1:A2)" in (r.get("product") or "") for r in rows)


def test_csv_sanitization_can_be_disabled_explicitly():
    data = _hostile_run()
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "raw.csv"
        write_csv(path, data["records"], ServiceRecord, sanitize=False)
        text = path.read_text(encoding="utf-8")
        assert "'+SUM(A1:A2)" not in text and "+SUM(A1:A2)" in text


def test_firewall_rule_csv_descriptions_are_safe():
    rule = FirewallRule(
        rule_id="IN-0001",
        description='=HYPERLINK("http://evil")',
        purpose="+SUM(A1:A2)",
        source="192.168.1.50",
    )
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "rules.csv"
        write_csv(path, [rule], FirewallRule)
        with path.open(newline="", encoding="utf-8") as handle:
            row = next(csv.DictReader(handle))
        assert row["description"].startswith("'=")
        assert row["purpose"].startswith("'+")
        assert row["source"] == "192.168.1.50"  # IPs untouched


def test_excel_cells_are_not_formulas():
    try:
        import openpyxl
    except ImportError:  # pragma: no cover - environment without openpyxl
        return
    from nmap_flow_analyzer.excel_export import export_excel

    data = _hostile_run()
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "out.xlsx"
        message = export_excel(
            path, data["metadata"], data["hosts"], data["records"],
            data["flows"], data["inbound"], data["outbound"],
            [], [], ["assumption"],
        )
        assert message is None
        workbook = openpyxl.load_workbook(path)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    assert cell.data_type != "f", (
                        f"{sheet.title}!{cell.coordinate} became a formula: "
                        f"{cell.value!r}"
                    )
        # Hostile hostname, product, NSE output, and command line all present
        # somewhere as literal text.
        all_text = " ".join(
            str(c.value)
            for ws in workbook.worksheets
            for row in ws.iter_rows()
            for c in row
            if c.value is not None
        )
        assert "HYPERLINK" in all_text
        assert "+SUM(A1:A2)" in all_text


def test_dot_output_cannot_be_injected():
    data = _hostile_run()
    dot = build_dot(
        data["hosts"], data["flows"], data["metadata"], base_config(), "summary"
    )
    # The raw injection string must never appear unescaped.
    assert 'host"; A -> B; "' not in dot
    # Every label remains inside its quoted string: strip escaped quotes and
    # the remaining quotes must be balanced per line.
    for line in dot.splitlines():
        stripped = line.replace('\\"', "")
        assert stripped.count('"') % 2 == 0, line


def test_mermaid_output_cannot_be_injected():
    data = _hostile_run()
    mmd = build_mermaid(
        data["hosts"], data["flows"], data["metadata"], base_config(), "summary"
    )
    assert "node] --> attacker" not in mmd
    assert "<script>" not in mmd


def test_html_report_escapes_untrusted_values():
    data = _hostile_run()
    reviews = [
        ManualReviewItem(
            item_id="MR-0001", category="test",
            finding="<script>alert(1)</script>",
            reason='host"; A -> B; "',
        )
    ]
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "report.html"
        write_html_report(
            path, data["metadata"], data["hosts"], data["records"],
            data["flows"], data["inbound"], data["outbound"], reviews, [],
            ["assumption"],
        )
        html_text = path.read_text(encoding="utf-8")
    # Raw markup never survives; the escaped literal text does.
    assert "<script>alert(1)</script>" not in html_text
    assert "&lt;script&gt;alert(1)&lt;/script&gt;" in html_text
    # The hostile hostname from the scan is also escaped, not dropped.
    assert 'host&quot;; A -&gt; B; &quot;' in html_text


def test_markdown_summary_escapes_untrusted_values():
    data = _hostile_run()
    reviews = [
        ManualReviewItem(
            item_id="MR-0001",
            category="test",
            finding="value|new-column",
            reason="<script>alert(1)</script>",
        )
    ]
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "summary.md"
        write_findings_summary(
            path, data["metadata"], data["hosts"], data["records"],
            data["flows"], data["inbound"], data["outbound"], reviews, [],
            ["assumption"],
        )
        md_text = path.read_text(encoding="utf-8")
    assert "value\\|new-column" in md_text
    assert "<script>" not in md_text
